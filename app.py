import streamlit as st
import anthropic
import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
import re
import io
from dotenv import load_dotenv

load_dotenv()

st.set_page_config(
    page_title="Transit Assist by ИС «KEDEN»",
    page_icon="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAIAAAD8GO2jAAAH6UlEQVR4nE2W349dVRXHv2vtfe+5c+fOTOdOO23pj2mhP2mxRQrKD6klTWkJwoOGxESD4dEXE/8MY4wPajQKUYNEMYiiBITWotTQCqWFltZSCm3pTGemt3PvnXt+n7PX8mHfO/Xp7HP22uvnPmt96L33TjAxERGTtVZE+i9EAJiJmAEwEQAAxExEftcwE5GqEhOzUREiBuB3RYWZmYm9BBE55whekYKUDQMEBQGquqQXqioKQESdCJhUoQqAVBUEhXp5EWEnDgQRUZG+mwRFXw+gIL+AAv6h3kcFCASCQFVVpa/av3kbAIsIDZQ5ESIABEAFUIiqqrJh0CCy/nnxyVFVEQGgoiLaF/F+qCrAzKTweSOoiqpf+IwbY0REXN8tIiZmZhZREQdVhRIT1AfmDfbPMhGBmABVEXUqQszixJVlPwkEcc6wAUFEoSoizpUqwsywFVOtAuqcw6A0A0uAd5LAPi4iElVxjpm9l6ICVRCpCqDkK6GqToSIRc/+8tmP3jxsq0E1CKR06oNypYpzTlR8dtWCIE769wxwrmQ1UMfMIo6I2RhiZjYgBpStlSK//Ps/1p3MnfzgxNzNrQf2NdeuhQiIiA1UyiRR+BrBQslYAyJjjKlUiNkY4yvhnFNXpnEM58okMSARB0Vy8rTGWfUfx4Y7nc6e3Wei3rq7djCbNM9NkppGfePDX3FZBmZALTGlcZxFcRaGZRS5MMp7oc2zorNYyfMijEyWS6ebrF1VGRvVvCiTbPWKSXrrWDYzu2x4qHrseK/Tnb52HWnqsmzZ+jVUr2f3fLFaDZw4w8Yq1FarBKpUqzoyYqGSFyTiklTLUqKkLLL0wicA7LETuHadgiDO8jqbyvCQYx6ZmBi6cMlc/jwYrucL7fb8Dd6/N55vBVPrSURU7OJCJ4/jIo41y/IwyqNQ4tTFEdI8j6Kaqha5JGnzgfsbIyPR9dlqo1E6lxGRNUIcZsn4jh2fheFsFD1orH3xpSSMetdnJ3dsz2+0yLANakNBEHCzaa2tVCvGVkQ1iaI0jBZvzPfm5tyNVvTpZ/PPPmeJOS8CololMGXJUSJhWN+w7sLOHb/51zu97mL0yIMHV05G7W7r4kUaG5navStf7FlbNaCKMUZVFhYW4oV2stBOWjc5jkyc5AvtCtBMs7GNG9nYelAzYZyd+69Yu+KJx0aWT+Sb73jpn//mNJtsjJz85PK+e/eY537dffBL4fRM99Jn2x5/zIoqQ4ssN4ZHRseWjY9XtgdEEHHGmCSKzx052lg+MXv8P72Z2WBsdPt9e3rvnixun1r+/e+uXrP25z/71YenPvjaoQOdbu/4qdNndmzd+/ihS91eumWzNeadH/zQ+vZJDAXSOGrPz+fdxbTV6s3OaS9009d53Rq6cuXqb1+oDw+bPffktXraC6sb1o+Pjh1/4+jf/vLq+nVrvvOtpz46/7G48lK7+8nKVcu2bd+55rbGjVZtwwarImSMb3DGVpqTk5XVqy3fqc6VZWGIzrz29+mr1+546uufv/ra6NZNxeVrCWR08yYneP7FlwTaHBv90Y9/enVmfrg+VK3V9mzduPzMmXocJkF14zNPWwVEhNkAGgzVep10fmam7C7mnU7YWih7vTHQqKjbsnHq6W+PNJu9I29Lrbbmnt0v//mvH3/6aRDUrs7fXLV61f6H7m+umHjl9cPt1vz3Dh7ofHB2yzNPF4s9q74pExGRiNYbI0EQGGZmw6AsiZIoItWrb719M0tvb06cuzZdWXdbu1J9+eVXKtXqQw98eVO91pyevvnC76Inn4iLcmZ2/kIvGhsZLlWg6mvghwhExVjDdqjTaoWtmy4MyzDMF8NKWQ7bSqZonT2bJtFt+/f96fDRmWvTjzy6f28an3v2udRYKovg/VN37tz1xutvHjnx7pNrV6ad7lCjYZnZlaVCVWGY/VwdG2+ONZtZnBRxnHa7yULbtTvtK5fHx8aX37fn5u7dp5//w9a7d+1bMXH+J7+wjQYZY7McWbZt7arevXdPTIznaRq3WiMTTWuMYTZQJSY/VImghCROeq1W3l3UNI3bCxWFKYqZslh16NHw4sVv7NzaWDmJ67O5uBqbvLs4eejR4W1bohs3vvmF7Q6ENK01GqSg999/V0W13/oHfR/ExE5cr7fosjyPYpelQemyoqgND1crZnh0WXt27uNTp4fzvHP4rfFdd43vvHNsat3U3ofLvGDDDqrOuaKgMx+eJiLtzyNvSwbQIJ4kjK0Ya2AME5w4ZqMqlWotmZs7/eaRQKReqTRWLJ86eKCMYj+ssjT1s5rOnz/rJ5qKigqBFLfmOMhPcyEvA2ViBYwxKmKCILD2yjvH8zDcdPCA5AX1b6PkeS4izGT9MO3PYPXMxZ5PlP3gJ1GnCjbG8wETqSobY1SdyOZ9XxUVVxTKVBal94kAY5iIrYoAJOJuMZWoyK1ieBoTlSLPjTHoI4gPCJZZihIEJi5cASiBwSBinwzLzDrglgG+gdmHAFVlpgF3eRkwL4GkirjCQx2BiDzRAPCtgcDWQ9qAKfsqABqQhMdGeMP9XQ906F8EQIn6YgAxL1EVAbAE+j9qYxCYvO99pf7ueqUDOtU+qA20DJzweAoiHRCi2sEhH4T3a8DG/Wj6EfgQAR8WlkjSyxMt/ae3flgAltnfQO7zHrPfZrCqcv/z0kksnfTLgaWlLb+QJdr/H1IEO0axQo+DAAAAAElFTkSuQmCC",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

* { font-family: 'Inter', sans-serif; box-sizing: border-box; }

/* Убираем стандартный streamlit padding */
.stApp { background: #F0F2F5; }
.block-container { padding: 0 !important; max-width: 100% !important; }
[data-testid="stSidebar"] { display: none; }
header[data-testid="stHeader"] { display: none; }
.stDeployButton { display: none; }

/* ── ШАПКА ── */
.keden-header {
    background: #ffffff;
    border-bottom: 1px solid #E0E4EA;
    padding: 0 2rem;
    height: 64px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    position: sticky;
    top: 0;
    z-index: 100;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
}
.keden-logo {
    display: flex;
    align-items: center;
    gap: 10px;
    font-size: 1.15rem;
    font-weight: 700;
    color: #1B4F8A;
    text-decoration: none;
}
.keden-logo-icon {
    width: 36px; height: 36px;
    background: #1B4F8A;
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    color: white; font-size: 1.1rem;
}
.keden-nav {
    display: flex; align-items: center; gap: 0.5rem;
}
.keden-nav a {
    color: #444; font-size: 0.875rem; text-decoration: none;
    padding: 0.4rem 0.8rem; border-radius: 4px;
    transition: background 0.15s;
}
.keden-nav a:hover { background: #F0F2F5; }

/* ── HERO БАННЕР ── */
.keden-hero {
    background: linear-gradient(135deg, #1B4F8A 0%, #1a6cb5 100%);
    color: white;
    padding: 2.5rem 2.5rem 2rem 2.5rem;
    margin-bottom: 0;
}
.keden-hero h1 {
    font-size: 1.6rem; font-weight: 700;
    margin: 0 0 0.4rem 0; letter-spacing: 0.02em;
    color: white !important;
}
.keden-hero p {
    font-size: 0.95rem; opacity: 0.85; margin: 0;
    color: white !important;
}
.hero-features {
    display: flex; gap: 2rem; margin-top: 1.5rem;
    border-top: 1px solid rgba(255,255,255,0.2);
    padding-top: 1.2rem;
}
.hero-feature { flex: 1; }
.hero-feature-title { font-weight: 600; font-size: 0.9rem; margin-bottom: 0.25rem; }
.hero-feature-desc { font-size: 0.8rem; opacity: 0.75; line-height: 1.4; }

/* ── ОСНОВНОЙ КОНТЕНТ ── */
.keden-main {
    display: flex;
    min-height: calc(100vh - 64px - 160px);
}

/* ── ЛЕВОЕ МЕНЮ ── */
.keden-sidebar {
    width: 260px;
    min-width: 260px;
    background: #ffffff;
    border-right: 1px solid #E0E4EA;
    padding: 1.2rem 0;
}
.sidebar-section-title {
    font-size: 0.7rem; font-weight: 600; color: #8A9BB5;
    text-transform: uppercase; letter-spacing: 0.1em;
    padding: 0.6rem 1.2rem 0.4rem;
}
.sidebar-item {
    display: flex; align-items: center; gap: 0.6rem;
    padding: 0.55rem 1.2rem;
    font-size: 0.875rem; color: #2C3E50; cursor: pointer;
    border-left: 3px solid transparent;
    transition: all 0.15s;
}
.sidebar-item:hover { background: #F5F7FA; }
.sidebar-item.active {
    background: #EBF3FF;
    border-left-color: #1B5EA8;
    color: #1B5EA8;
    font-weight: 500;
}
.sidebar-item .item-icon { font-size: 1rem; width: 20px; text-align: center; }

/* ── КОНТЕНТ ОБЛАСТЬ ── */
.keden-content {
    flex: 1;
    padding: 1.5rem 2rem;
    overflow: auto;
}

/* ── КАРТОЧКИ ── */
.keden-card {
    background: #ffffff;
    border: 1px solid #E0E4EA;
    border-radius: 8px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.keden-card-title {
    font-size: 1rem; font-weight: 600; color: #1B2D45;
    margin-bottom: 1rem;
    padding-bottom: 0.75rem;
    border-bottom: 1px solid #F0F2F5;
    display: flex; align-items: center; gap: 0.5rem;
}
.keden-card-title .card-icon {
    width: 28px; height: 28px;
    background: #EBF3FF; border-radius: 6px;
    display: flex; align-items: center; justify-content: center;
    font-size: 0.85rem;
}

/* ── ПОЛЯ ДАННЫХ (как в KEDEN) ── */
.field-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; }
.field-grid-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 1rem; }
.field-group { display: flex; flex-direction: column; gap: 0.3rem; }
.field-label { font-size: 0.78rem; color: #6B7C93; font-weight: 500; }
.field-value {
    background: #F5F7FA;
    border: 1px solid #E0E4EA;
    border-radius: 6px;
    padding: 0.55rem 0.85rem;
    font-size: 0.875rem; color: #1B2D45;
    min-height: 38px;
}
.field-value.highlight {
    background: #EBF3FF;
    border-color: #B8D4F0;
    color: #1B5EA8;
    font-weight: 600;
}

/* ── КНОПКИ ── */
.stButton > button {
    background: #1B5EA8 !important;
    color: white !important;
    border: none !important;
    border-radius: 6px !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    padding: 0.55rem 1.4rem !important;
    height: auto !important;
    transition: background 0.15s !important;
    letter-spacing: 0.01em !important;
}
.stButton > button:hover { background: #1a4d8f !important; }
.stButton > button[kind="secondary"] {
    background: white !important;
    color: #1B5EA8 !important;
    border: 1px solid #1B5EA8 !important;
}

/* ── ЗАГРУЗЧИК ФАЙЛОВ ── */
[data-testid="stFileUploader"] {
    background: #F5F7FA;
    border: 2px dashed #C5D5E8;
    border-radius: 8px;
    padding: 1rem;
}
[data-testid="stFileUploader"]:hover { border-color: #1B5EA8; }

/* ── RADIO КНОПКИ (режим) ── */
.stRadio > div { gap: 0 !important; }
.stRadio label {
    background: white !important;
    border: 1px solid #E0E4EA !important;
    border-radius: 6px !important;
    padding: 0.75rem 1rem !important;
    margin-bottom: 0.5rem !important;
    cursor: pointer !important;
    transition: all 0.15s !important;
    display: flex !important;
}
.stRadio label:hover { border-color: #1B5EA8 !important; background: #F5F8FF !important; }

/* ── СТАТУСНЫЕ БЛОКИ ── */
.keden-alert {
    border-radius: 6px;
    padding: 0.85rem 1.1rem;
    margin: 0.6rem 0;
    font-size: 0.875rem;
    display: flex; align-items: flex-start; gap: 0.6rem;
    line-height: 1.5;
}
.keden-alert-icon { font-size: 1rem; flex-shrink: 0; margin-top: 0.05rem; }
.keden-alert.success { background: #F0FBF4; border: 1px solid #B2DFC2; color: #1E5E34; }
.keden-alert.error   { background: #FEF3F2; border: 1px solid #F5C6C3; color: #92180F; }
.keden-alert.warning { background: #FFFBF0; border: 1px solid #F5DFA0; color: #7A5200; }
.keden-alert.info    { background: #EBF3FF; border: 1px solid #B8D4F0; color: #1B4F8A; }

/* ── МЕТРИКИ ── */
.keden-metrics {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(130px, 1fr));
    gap: 0.75rem;
    margin: 1rem 0;
}
.keden-metric {
    background: white;
    border: 1px solid #E0E4EA;
    border-radius: 8px;
    padding: 1rem;
    text-align: center;
}
.keden-metric-value {
    font-size: 1.5rem; font-weight: 700; color: #1B5EA8;
    line-height: 1.2;
}
.keden-metric-label {
    font-size: 0.72rem; color: #6B7C93; margin-top: 0.3rem;
    font-weight: 500; text-transform: uppercase; letter-spacing: 0.06em;
}

/* ── СТРОКИ ПРОВЕРКИ ── */
.check-item {
    display: flex; align-items: flex-start; gap: 0.75rem;
    padding: 0.85rem 1rem;
    border-radius: 6px;
    border: 1px solid #E0E4EA;
    background: white;
    margin: 0.4rem 0;
}
.check-item.ok      { border-left: 4px solid #27AE60; }
.check-item.error   { border-left: 4px solid #E74C3C; }
.check-item.warning { border-left: 4px solid #F39C12; }
.check-item.info    { border-left: 4px solid #1B5EA8; }
.check-icon { font-size: 1.1rem; flex-shrink: 0; }
.check-body { flex: 1; }
.check-title { font-weight: 600; font-size: 0.875rem; color: #1B2D45; margin-bottom: 0.2rem; }
.check-detail { font-size: 0.8rem; color: #6B7C93; line-height: 1.5; }

/* ── СЕКЦИЯ ── */
.section-header {
    font-size: 0.8rem; font-weight: 600; color: #1B5EA8;
    text-transform: uppercase; letter-spacing: 0.1em;
    margin: 1.4rem 0 0.7rem 0;
    padding-bottom: 0.4rem;
    border-bottom: 2px solid #EBF3FF;
    display: flex; align-items: center; gap: 0.5rem;
}

/* ── РИСК БЕЙДЖИ ── */
.risk-badge {
    display: inline-block;
    padding: 0.15rem 0.6rem;
    border-radius: 12px;
    font-size: 0.72rem;
    font-weight: 600;
}
.risk-HIGH   { background: #FDECEA; color: #C0392B; }
.risk-MED    { background: #FEF4E4; color: #E67E22; }
.risk-LOW    { background: #FEFEE4; color: #9B8800; }
.risk-NONE   { background: #EAFAF1; color: #1E8449; }
.risk-UNKN   { background: #F2F3F4; color: #566573; }

/* ── ТАБЛИЦА ПОЗИЦИЙ ── */
.hs-table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
.hs-table th {
    background: #1B4F8A; color: white;
    padding: 0.6rem 0.75rem;
    text-align: left; font-weight: 500;
    font-size: 0.75rem;
}
.hs-table td {
    padding: 0.55rem 0.75rem;
    border-bottom: 1px solid #F0F2F5;
    color: #2C3E50;
    vertical-align: middle;
}
.hs-table tr:nth-child(even) td { background: #F9FAFB; }
.hs-table tr:hover td { background: #EBF3FF; }
.hs-table .total-row td { background: #EBF3FF !important; font-weight: 700; color: #1B4F8A; }
.hs-code { font-family: monospace; font-weight: 600; color: #1B5EA8; }

/* ── ФУТЕР ── */
.keden-footer {
    background: #1B2D45;
    color: rgba(255,255,255,0.7);
    padding: 1.5rem 2.5rem;
    font-size: 0.8rem;
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-top: auto;
}
.footer-logo { color: white; font-weight: 600; font-size: 0.9rem; }
</style>
""", unsafe_allow_html=True)

# ── Системный промпт ───────────────────────────────────────────────────────────
SYSTEM_PROMPT = """Ты — эксперт по таможенному декларированию Республики Казахстан и обработке коммерческих инвойсов для подготовки данных под ИС KEDEN при оформлении ТРАНЗИТНЫХ ДЕКЛАРАЦИЙ (процедура таможенного транзита, глава 22 ТК ЕАЭС).
Декларации для выпуска в свободное обращение в РК НЕ оформляются — товар следует транзитом до места назначения.

Возможные сценарии транзита:
А) Транзит через РК в страну ЕАЭС (например: Китай → РК → РФ / Беларусь / Армения / Кыргызстан) — казахстанская таможня требует полный пакет разрешительных документов.
Б) Транзит через ЕАЭС в третью страну (например: Китай → РК → Узбекистан / Таджикистан / Афганистан) — сокращённый пакет, только запреты и санитарный/ветеринарный/фитосанитарный контроль.

Выполни следующее без лишних пояснений, сразу результат:

1. SANITY CHECKS — обязательные проверки:
• Σ Net Weight ≤ Σ Gross Weight — если нарушено, КРИТИЧНО
• Σ (Qty × Unit Price) = Σ Total Value (допуск 0.5%) — если нарушено, КРИТИЧНО
• Σ PKGS по строкам = TOTAL PKGS из шапки — если нарушено, КРИТИЧНО
• По строкам: Net ≤ Gross, Qty×Price = Total, веса > 0, цена > 0

2. ОЧИСТКА ДАННЫХ:
• Удали строки TOTAL/Итого, пустые строки, строки без кода ТН ВЭД
• Наименования товаров — ТОЛЬКО на русском языке как есть в инвойсе. НЕ переводить на английский, НЕ придумывать названия
• Пересчитай Total Value = Qty × Unit Price
• Если Net Weight пустой — рассчитай как Gross Weight / 1.15

3. ЛОГИКА PKGS — ВАЖНО:
Данные инвойса уже предобработаны: объединённые ячейки разъединены, логика PKGS применена Python-кодом.
Твоя задача при группировке:
• PKGS = 0 означает что товар находится в той же упаковке что и другой товар — при суммировании по HS6 прибавляй 0
• PKGS = 1 означает что этот товар является «представителем» своей упаковки — при суммировании по HS6 прибавляй 1
• При группировке по HS6 — суммируй PKGS как есть, НЕ пересчитывай
• Контроль: итоговая сумма PKGS должна равняться количеству коробок в строке TOTAL инвойса

4. ГРУППИРОВКА по первым 6 цифрам кода ТН ВЭД:
• Объединяй товары только по первым 6 цифрам
• При объединении суммируй: PCS, вес нетто, вес брутто, стоимость, PKGS
• Наименование — перечисли все уникальные через запятую (только русский язык)
• Полные коды HS10 — сохрани для справки

5. ОПРЕДЕЛЕНИЕ СЦЕНАРИЯ ТРАНЗИТА:
• Сценарий А: получатель в стране ЕАЭС (РФ, РБ, РК, РА, КР)
• Сценарий Б: получатель в третьей стране (вне ЕАЭС)

6. ПРОВЕРКА РАЗРЕШИТЕЛЬНЫХ ДОКУМЕНТОВ:

СЦЕНАРИЙ А (полный пакет):
• Решения Коллегии ЕЭК № 30, № 134 — запреты и ограничения
• ТР ТС/ТР ЕАЭС — сертификат/декларация соответствия
• Решение КТС № 299 — СГР
• Решение КТС № 317 — ветеринарный сертификат
• Решение Совета ЕЭК № 157 — фитосанитарный сертификат
• Нотификация ФСБ — для IT, шифрование, Wi-Fi, Bluetooth
• Сертификат происхождения (СТ-1, Form A, EAV)
• Лицензии Минпромторг/ФСТЭК/ФСВТС — двойное назначение
• СИТЕС, озоноразрушающие, драгметаллы, прекурсоры

СЦЕНАРИЙ Б (сокращённый пакет):
• Только запреты (Приложение № 1 к Решению № 30)
• Санитарный/ветеринарный/фитосанитарный контроль
• СИТЕС, экспортный контроль
• ТР ТС, СГР, нотификация ФСБ — НЕ требуются, пометить «Не требуется при транзите в третью страну»

Уровень риска: ВЫСОКИЙ / СРЕДНИЙ / НИЗКИЙ / НЕТ / ТРЕБУЕТ УТОЧНЕНИЯ

7. СВЕРКА МЕЖДУ ДОКУМЕНТАМИ (если загружено несколько файлов):
• Сравни отправителя, получателя, номер инвойса, общий вес брутто, количество мест

Верни ТОЛЬКО валидный JSON без пояснений и markdown:
{
  "sanity_checks": {"passed": true, "issues": []},
  "statistics": {
    "invoice_number": "", "invoice_date": "", "sender": "", "receiver": "",
    "country_from": "", "country_to": "", "transit_scenario": "", "transit_scenario_description": "",
    "doc_type": "", "rows_original": 0, "rows_grouped": 0,
    "total_qty": 0, "total_pkgs": 0, "total_net_weight": 0, "total_gross_weight": 0,
    "total_value": 0, "currency": ""
  },
  "grouped_items": [{
    "num": 1, "hs6": "", "hs10_codes": "", "name_ru": "",
    "qty": 0, "pkgs": 0, "net_weight": 0, "gross_weight": 0, "value": 0,
    "source_rows": 0, "note": ""
  }],
  "permit_documents": [{
    "num": 1, "hs6": "", "name": "", "scenario": "", "applicable_act": "",
    "restriction_category": "", "recommended_doc": "", "issuing_authority": "",
    "risk_level": "", "comment": ""
  }],
  "risks": [{
    "num": 1, "problem_type": "", "source": "", "position": "",
    "description": "", "actual_value": "", "expected_value": "",
    "criticality": "", "recommendation": ""
  }],
  "cross_check": {
    "documents_compared": [],
    "discrepancies": [{"field": "", "doc1_value": "", "doc2_value": "", "criticality": "", "comment": ""}],
    "matches": []
  }
}"""


def extract_text_from_pdf(file_bytes):
    parts = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t: parts.append(t)
            for table in page.extract_tables():
                for row in table:
                    if row: parts.append(" | ".join(str(c) if c else "" for c in row))
    return "\n".join(parts)

def extract_text_from_excel(file_bytes):
    """
    Извлечение данных из Excel с правильной логикой PKGS согласно промпту:
    1. До разъединения сохраняем группы строк объединённых ячеек столбца PKGS
    2. Разъединяем все объединённые ячейки, копируя значения
    3. Применяем логику PKGS:
       - Внутри каждой объединённой упаковки: PKGS=1 → товару с max брутто,
         остальным PKGS=0
       - Если несколько товаров с одинаковым max брутто → первому из них
       - Одиночные строки (не объединённые) → оставляем как есть
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    parts = []
    SKIP_KEYWORDS = ['инструк', 'справ', 'шаблон', 'template', 'help', 'readme']

    for sn in wb.sheetnames:
        ws = wb[sn]
        if any(kw in sn.lower() for kw in SKIP_KEYWORDS):
            continue

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # ── ШАГ 1: Находим заголовки ─────────────────────────────────────────────
        header_row = None
        pkgs_col = None
        gross_col = None

        for r in range(1, min(30, max_row + 1)):
            row_str = " ".join(
                str(ws.cell(r, c).value).lower()
                for c in range(1, max_col + 1)
                if ws.cell(r, c).value
            )
            if any(kw in row_str for kw in ['pkgs', 'кол-во мест', 'мест', 'package']):
                header_row = r
                for c in range(1, max_col + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    v_low = str(v).lower()
                    if any(kw in v_low for kw in ['pkgs', 'кол-во мест', 'мест', 'package']):
                        pkgs_col = c
                    if any(kw in v_low for kw in ['брутто', 'gross weight', 'gross']):
                        gross_col = c
                break

        # ── ШАГ 2: Сохраняем группы объединённых ячеек столбца PKGS ─────────────
        # Делаем ЭТО ДО разъединения — пока ещё знаем какие строки объединены
        merged_ranges = list(ws.merged_cells.ranges)
        pkgs_merged_groups = {}  # {min_row: [список строк диапазона]}
        pkgs_merged_rows = set()

        if pkgs_col:
            for mr in merged_ranges:
                if mr.min_col == pkgs_col and mr.max_col == pkgs_col:
                    rows = list(range(mr.min_row, mr.max_row + 1))
                    pkgs_merged_groups[mr.min_row] = rows
                    for r in rows:
                        pkgs_merged_rows.add(r)

        # ── ШАГ 3: Сохраняем все значения объединённых ячеек ─────────────────────
        merge_fill = {}
        for mr in merged_ranges:
            mv = ws.cell(mr.min_row, mr.min_col).value
            for ri in range(mr.min_row, mr.max_row + 1):
                for ci in range(mr.min_col, mr.max_col + 1):
                    merge_fill[(ri, ci)] = mv

        # ── ШАГ 4: Разъединяем и проставляем значения ────────────────────────────
        for mr in merged_ranges:
            ws.unmerge_cells(str(mr))
        for (ri, ci), v in merge_fill.items():
            ws.cell(ri, ci).value = v

        # ── ШАГ 5: Применяем логику PKGS для объединённых групп ──────────────────
        if pkgs_col and gross_col and pkgs_merged_groups:
            for start_row, row_indices in pkgs_merged_groups.items():
                # Пропускаем строки заголовка и шапки
                data_rows = [ri for ri in row_indices if ri > (header_row or 0)]
                if not data_rows:
                    continue

                if len(data_rows) == 1:
                    # Один товар — PKGS=1
                    ws.cell(data_rows[0], pkgs_col).value = 1
                else:
                    # Несколько товаров — max брутто получает PKGS=1, остальные 0
                    gross_values = []
                    for ri in data_rows:
                        gv = ws.cell(ri, gross_col).value
                        try:
                            gross_values.append(float(gv) if gv is not None else 0.0)
                        except (ValueError, TypeError):
                            gross_values.append(0.0)

                    max_gross = max(gross_values) if gross_values else 0.0
                    assigned = False
                    for ri, gv in zip(data_rows, gross_values):
                        if not assigned and gv == max_gross:
                            ws.cell(ri, pkgs_col).value = 1
                            assigned = True
                        else:
                            # Именно 0, не пустое
                            ws.cell(ri, pkgs_col).value = 0

        # ── ШАГ 6: Формируем текстовый вывод ─────────────────────────────────────
        sheet_rows = []
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                val = ws.cell(r, c).value
                if val is None:
                    row_vals.append("")
                elif isinstance(val, float):
                    row_vals.append(
                        str(int(val)) if val == int(val) else str(round(val, 4))
                    )
                else:
                    s = str(val).strip()
                    if len(s) > 500:
                        s = s[:500]
                    row_vals.append(s)

            if not any(v for v in row_vals):
                continue

            sheet_rows.append(" | ".join(row_vals).rstrip(" |"))

        if sheet_rows:
            parts.append(f"=== Лист: {sn} ===")
            prev = None
            for row_str in sheet_rows:
                if row_str != prev:
                    parts.append(row_str)
                    prev = row_str

    result = "\n".join(parts)

    MAX_CHARS = 120000
    if len(result) > MAX_CHARS:
        result = result[:MAX_CHARS]
        result += "\n\n[ПРИМЕЧАНИЕ: Данные обрезаны из-за размера файла. Обработаны первые строки.]"

    return result
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    parts = []

    SKIP_KEYWORDS = ['инструк', 'справ', 'шаблон', 'template', 'help', 'readme']

    for sn in wb.sheetnames:
        ws = wb[sn]

        if any(kw in sn.lower() for kw in SKIP_KEYWORDS):
            continue

        # ── ШАГ 1: Сохраняем значения объединённых ячеек ────────────────────────
        merge_fill = {}
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            merge_value = ws.cell(merged_range.min_row, merged_range.min_col).value
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    merge_fill[(row_idx, col_idx)] = merge_value

        # ── ШАГ 2: Разъединяем и проставляем значения ───────────────────────────
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
        for (row_idx, col_idx), value in merge_fill.items():
            ws.cell(row_idx, col_idx).value = value

        # ── ШАГ 3: Читаем все данные в массив ───────────────────────────────────
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Читаем данные в список словарей {col_idx: value}
        all_rows = []
        header_row = None
        pkgs_col = None
        gross_col = None

        for r in range(1, max_row + 1):
            row_data = {}
            for c in range(1, max_col + 1):
                row_data[c] = ws.cell(r, c).value

            # Ищем строку заголовков по ключевым словам
            if header_row is None:
                row_str = " ".join(str(v).lower() for v in row_data.values() if v)
                if any(kw in row_str for kw in ['pkgs', 'кол-во мест', 'мест', 'package']):
                    header_row = r
                    # Определяем номера столбцов PKGS и брутто
                    for c, v in row_data.items():
                        if v is None:
                            continue
                        v_low = str(v).lower()
                        if any(kw in v_low for kw in ['pkgs', 'кол-во мест', 'мест', 'package']):
                            pkgs_col = c
                        if any(kw in v_low for kw in ['брутто', 'gross', 'gross weight']):
                            gross_col = c

            all_rows.append((r, row_data))

        # ── ШАГ 4: Применяем логику PKGS ─────────────────────────────────────────
        # Только если нашли столбцы PKGS и брутто
        if pkgs_col and gross_col and header_row:
            # Группируем строки по значению PKGS (номеру упаковки)
            # Строки где PKGS не None и > 0 — начало упаковки
            # Строки где PKGS = None — продолжение той же упаковки

            current_pkg_rows = []   # строки текущей упаковки
            current_pkg_val = None  # значение PKGS текущей упаковки
            pkgs_groups = []        # список групп [(pkg_value, [row_indices])]

            for r, row_data in all_rows:
                if r <= header_row:
                    continue
                # Проверяем есть ли данные в строке
                if not any(v for v in row_data.values() if v is not None):
                    continue

                pkgs_val = row_data.get(pkgs_col)

                if pkgs_val is not None and pkgs_val != 0 and pkgs_val != "":
                    # Новая упаковка начинается
                    if current_pkg_rows:
                        pkgs_groups.append((current_pkg_val, current_pkg_rows))
                    current_pkg_val = pkgs_val
                    current_pkg_rows = [r]
                elif current_pkg_rows:
                    # Продолжение текущей упаковки
                    current_pkg_rows.append(r)
                else:
                    # Строка до первой упаковки — пропускаем
                    pass

            # Добавляем последнюю группу
            if current_pkg_rows:
                pkgs_groups.append((current_pkg_val, current_pkg_rows))

            # Для каждой группы применяем логику: max брутто → PKGS=1, остальные → PKGS=0
            for pkg_val, row_indices in pkgs_groups:
                if len(row_indices) == 1:
                    # Один товар в упаковке — просто PKGS=1
                    ws.cell(row_indices[0], pkgs_col).value = 1
                else:
                    # Несколько товаров — ищем max брутто
                    gross_values = []
                    for ri in row_indices:
                        gv = ws.cell(ri, gross_col).value
                        try:
                            gross_values.append(float(gv) if gv is not None else 0.0)
                        except (ValueError, TypeError):
                            gross_values.append(0.0)

                    max_gross = max(gross_values)
                    assigned = False
                    for ri, gv in zip(row_indices, gross_values):
                        if not assigned and gv == max_gross:
                            # Первый с максимальным весом → PKGS=1
                            ws.cell(ri, pkgs_col).value = 1
                            assigned = True
                        else:
                            # Остальные → PKGS=0 (именно ноль, не пустое)
                            ws.cell(ri, pkgs_col).value = 0

        # ── ШАГ 5: Формируем текстовый вывод ────────────────────────────────────
        sheet_rows = []
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                val = ws.cell(r, c).value
                if val is None:
                    row_vals.append("")
                elif isinstance(val, float):
                    row_vals.append(str(int(val)) if val == int(val) else str(round(val, 4)))
                else:
                    s = str(val).strip()
                    if len(s) > 500:
                        s = s[:500]
                    row_vals.append(s)

            if not any(v for v in row_vals):
                continue

            sheet_rows.append(" | ".join(row_vals).rstrip(" |"))

        if sheet_rows:
            parts.append(f"=== Лист: {sn} ===")
            prev = None
            for row_str in sheet_rows:
                if row_str != prev:
                    parts.append(row_str)
                    prev = row_str

    result = "\n".join(parts)

    MAX_CHARS = 120000
    if len(result) > MAX_CHARS:
        result = result[:MAX_CHARS]
        result += "\n\n[ПРИМЕЧАНИЕ: Данные обрезаны из-за размера файла. Обработаны первые строки.]"

    return result

def call_claude_api(text):
    key = os.getenv("ANTHROPIC_API_KEY")
    if not key: raise ValueError("API ключ не найден. Проверьте файл .env")
    client = anthropic.Anthropic(api_key=key)
    msg = client.messages.create(
        model="claude-sonnet-4-20250514", max_tokens=16000,
        system=SYSTEM_PROMPT,
        messages=[{"role":"user","content":f"Обработай инвойс, верни JSON:\n\n{text}"}]
    )
    resp = re.sub(r"```json\s*|```\s*","", msg.content[0].text).strip()
    try: return json.loads(resp)
    except:
        m = re.search(r"\{.*\}", resp, re.DOTALL)
        if m: return json.loads(m.group())
        raise ValueError(f"Ошибка разбора ответа:\n{resp[:400]}")

def create_excel_report(data, mode="full"):
    wb = openpyxl.Workbook()
    YEL = PatternFill("solid",fgColor="FFF3CD"); RED=PatternFill("solid",fgColor="F8D7DA")
    GRN = PatternFill("solid",fgColor="D4EDDA"); HDR=PatternFill("solid",fgColor="1B4F8A")
    TOT = PatternFill("solid",fgColor="EBF3FF"); BLU=PatternFill("solid",fgColor="EBF3FF")
    WB  = Font(name="Arial",bold=True,color="FFFFFF",size=10)
    KB  = Font(name="Arial",bold=True,size=10); NRM=Font(name="Arial",size=9)
    TF  = Font(name="Arial",bold=True,size=13,color="1B4F8A")
    TH  = Side(style="thin",color="CCCCCC"); BRD=Border(left=TH,right=TH,top=TH,bottom=TH)
    CTR = Alignment(horizontal="center",vertical="center",wrap_text=True)
    LFT = Alignment(horizontal="left",vertical="center",wrap_text=True)

    def hdr_row(ws,r,n):
        for c in range(1,n+1):
            cell=ws.cell(row=r,column=c); cell.fill=HDR; cell.font=WB; cell.alignment=CTR; cell.border=BRD

    stats=data.get("statistics",{}); sanity=data.get("sanity_checks",{})

    if mode=="full":
        ws1=wb.active; ws1.title="Статистика"
        ws1.column_dimensions["A"].width=35; ws1.column_dimensions["B"].width=50
        r=1
        if not sanity.get("passed",True):
            ws1.cell(row=r,column=1,value="⛔ КРИТИЧЕСКОЕ ПРЕДУПРЕЖДЕНИЕ")
            ws1.cell(row=r,column=1).font=Font(name="Arial",bold=True,size=12,color="FFFFFF")
            ws1.cell(row=r,column=1).fill=PatternFill("solid",fgColor="C62828")
            ws1.merge_cells(f"A{r}:B{r}"); ws1.cell(row=r,column=1).alignment=CTR; r+=1
            for iss in sanity.get("issues",[]):
                ws1.cell(row=r,column=1,value=iss).fill=RED; ws1.merge_cells(f"A{r}:B{r}")
                ws1.cell(row=r,column=1).font=Font(name="Arial",bold=True,size=9,color="C62828"); r+=1
            r+=1
        ws1.cell(row=r,column=1,value="Transit Assist by ИС «KEDEN» — Транзитная декларация").font=TF
        ws1.merge_cells(f"A{r}:B{r}"); r+=2
        for lbl,val in [
            ("Номер инвойса",stats.get("invoice_number","—")),("Дата инвойса",stats.get("invoice_date","—")),
            ("Отправитель",stats.get("sender","—")),("Получатель",stats.get("receiver","—")),
            ("Страна отправления",stats.get("country_from","—")),("Страна назначения",stats.get("country_to","—")),
            ("Сценарий транзита",f"{stats.get('transit_scenario','—')} — {stats.get('transit_scenario_description','—')}"),
            ("Тип документа",stats.get("doc_type","—")),("Строк в инвойсе",stats.get("rows_original","—")),
            ("Позиций после группировки",stats.get("rows_grouped","—")),("Итого количество (шт)",stats.get("total_qty","—")),
            ("Итого мест (PKGS)",stats.get("total_pkgs","—")),("Итого вес нетто (кг)",stats.get("total_net_weight","—")),
            ("Итого вес брутто (кг)",stats.get("total_gross_weight","—")),
            ("Итого стоимость",f"{stats.get('total_value','—')} {stats.get('currency','')}"),
        ]:
            ws1.cell(row=r,column=1,value=lbl).font=Font(name="Arial",bold=True,size=9)
            ws1.cell(row=r,column=1).fill=BLU; ws1.cell(row=r,column=1).border=BRD
            ws1.cell(row=r,column=2,value=str(val)).font=NRM
            ws1.cell(row=r,column=2).border=BRD; ws1.cell(row=r,column=2).alignment=LFT; r+=1
        ws2=wb.create_sheet("Сводная HS6 (KEDEN)")
    else:
        ws2=wb.active; ws2.title="Сводная HS6 (KEDEN)"

    h2=["№","Код ТН ВЭД\n(6 зн.)","Полные коды HS10","Наименование (RU)","Кол-во\n(шт)","Мест\n(PKGS)","Вес нетто\n(кг)","Вес брутто\n(кг)","Стоимость","Строк","Примечание"]
    cw=[5,12,22,40,10,8,12,12,14,7,28]
    for i,(h,w) in enumerate(zip(h2,cw),1): ws2.column_dimensions[get_column_letter(i)].width=w
    if mode=="quick":
        info=f"Инвойс: {stats.get('invoice_number','—')}  |  {stats.get('sender','—')} → {stats.get('receiver','—')}  |  Сценарий: {stats.get('transit_scenario','—')}"
        ws2.cell(row=1,column=1,value=info).font=Font(name="Arial",bold=True,size=9,color="1B4F8A")
        ws2.merge_cells(f"A1:{get_column_letter(len(h2))}1")
        tot=f"Итого: {stats.get('total_qty','—')} шт  |  {stats.get('total_pkgs','—')} мест  |  Нетто: {stats.get('total_net_weight','—')} кг  |  Брутто: {stats.get('total_gross_weight','—')} кг  |  {stats.get('total_value','—')} {stats.get('currency','')}"
        ws2.cell(row=2,column=1,value=tot).font=Font(name="Arial",size=9,color="444444")
        ws2.merge_cells(f"A2:{get_column_letter(len(h2))}2"); hr=3
    else: hr=1
    ws2.row_dimensions[hr].height=36
    for i,h in enumerate(h2,1): ws2.cell(row=hr,column=i,value=h)
    hdr_row(ws2,hr,len(h2)); ws2.freeze_panes=f"A{hr+1}"
    ws2.auto_filter.ref=f"A{hr}:{get_column_letter(len(h2))}{hr}"
    items=data.get("grouped_items",[]); ds=hr+1
    for idx,item in enumerate(items,ds):
        has_m=item.get("hs10_codes","") and "," in str(item.get("hs10_codes",""))
        fl=YEL if has_m else None
        for col,val in enumerate([item.get("num",idx-ds+1),item.get("hs6",""),item.get("hs10_codes",""),item.get("name_ru",""),item.get("qty",0),item.get("pkgs",0),item.get("net_weight",0),item.get("gross_weight",0),item.get("value",0),item.get("source_rows",1),item.get("note","")],1):
            cell=ws2.cell(row=idx,column=col,value=val); cell.font=NRM; cell.border=BRD; cell.alignment=LFT
            if fl: cell.fill=fl
    tr=ds+len(items); ws2.cell(row=tr,column=1,value="ИТОГО").font=KB
    ws2.cell(row=tr,column=1).fill=TOT; ws2.cell(row=tr,column=1).border=BRD; ws2.cell(row=tr,column=1).alignment=CTR
    for ci,cl in [(6,"F"),(7,"G"),(8,"H"),(9,"I"),(10,"J")]:
        c=ws2.cell(row=tr,column=ci,value=f"=SUM({cl}{ds}:{cl}{tr-1})")
        c.font=KB; c.fill=TOT; c.border=BRD; c.alignment=CTR
    for col in [2,3,4,5,11,12]:
        ws2.cell(row=tr,column=col).fill=TOT; ws2.cell(row=tr,column=col).border=BRD

    if mode=="full":
        ws3=wb.create_sheet("Разрешительные документы")
        ws3.cell(row=1,column=1,value=f"Сценарий: {stats.get('transit_scenario','—')} — {stats.get('transit_scenario_description','—')}").font=Font(name="Arial",bold=True,size=10,color="1B4F8A")
        ws3.merge_cells("A1:J1")
        h3=["№","Код ТН ВЭД","Наименование","Сценарий","Применимый акт","Категория","Рекомендуемый документ","Орган","Уровень риска","Комментарий"]
        cw3=[5,12,30,8,25,20,30,25,12,40]
        for i,(h,w) in enumerate(zip(h3,cw3),1): ws3.column_dimensions[get_column_letter(i)].width=w
        HR3=3; ws3.row_dimensions[HR3].height=36
        for i,h in enumerate(h3,1): ws3.cell(row=HR3,column=i,value=h)
        hdr_row(ws3,HR3,len(h3)); ws3.freeze_panes=f"A{HR3+1}"
        ws3.auto_filter.ref=f"A{HR3}:{get_column_letter(len(h3))}{HR3}"
        RF={"ВЫСОКИЙ":PatternFill("solid",fgColor="FDECEA"),"СРЕДНИЙ":PatternFill("solid",fgColor="FEF4E4"),
            "НИЗКИЙ":PatternFill("solid",fgColor="FEFEE4"),"НЕТ":PatternFill("solid",fgColor="EAFAF1"),
            "ТРЕБУЕТ УТОЧНЕНИЯ":PatternFill("solid",fgColor="F2F3F4")}
        for idx,doc in enumerate(data.get("permit_documents",[]),HR3+1):
            risk=doc.get("risk_level","НЕТ"); fl=RF.get(risk)
            for col,val in enumerate([doc.get("num",idx-HR3),doc.get("hs6",""),doc.get("name",""),doc.get("scenario",""),doc.get("applicable_act",""),doc.get("restriction_category",""),doc.get("recommended_doc",""),doc.get("issuing_authority",""),risk,doc.get("comment","")],1):
                cell=ws3.cell(row=idx,column=col,value=val); cell.font=NRM; cell.border=BRD; cell.alignment=LFT
                if fl: cell.fill=fl

    risks=data.get("risks",[])
    if mode=="full" or (mode=="quick" and risks):
        ws4=wb.create_sheet("Риски и замечания"); rhr=1
        if mode=="quick":
            ws4.cell(row=1,column=1,value=f"⚠️ Замечаний: {len(risks)}  |  Инвойс: {stats.get('invoice_number','—')}").font=Font(name="Arial",bold=True,size=10,color="C62828")
            ws4.cell(row=1,column=1).fill=PatternFill("solid",fgColor="FFF3CD"); ws4.merge_cells("A1:I1"); rhr=2
        h4=["№","Тип проблемы","Источник","Позиция","Описание","Факт. значение","Ожид. значение","Критичность","Рекомендация"]
        cw4=[5,20,12,15,45,20,20,12,40]
        for i,(h,w) in enumerate(zip(h4,cw4),1): ws4.column_dimensions[get_column_letter(i)].width=w
        ws4.row_dimensions[rhr].height=36
        for i,h in enumerate(h4,1): ws4.cell(row=rhr,column=i,value=h)
        hdr_row(ws4,rhr,len(h4)); ws4.freeze_panes=f"A{rhr+1}"; ws4.auto_filter.ref=f"A{rhr}:{get_column_letter(len(h4))}{rhr}"
        CF={"КРИТИЧНО":PatternFill("solid",fgColor="FDECEA"),"СРЕДНЕ":PatternFill("solid",fgColor="FEF4E4"),"ИНФО":PatternFill("solid",fgColor="FEFEE4")}
        if not risks:
            dr=rhr+1; ws4.cell(row=dr,column=1,value="Замечаний не выявлено").font=Font(name="Arial",bold=True,size=10,color="1E5E34")
            ws4.cell(row=dr,column=1).fill=GRN; ws4.merge_cells(f"A{dr}:{get_column_letter(len(h4))}{dr}"); ws4.cell(row=dr,column=1).alignment=CTR
        else:
            for idx,risk in enumerate(risks,rhr+1):
                crit=risk.get("criticality","ИНФО"); fl=CF.get(crit)
                for col,val in enumerate([risk.get("num",idx-rhr),risk.get("problem_type",""),risk.get("source",""),risk.get("position",""),risk.get("description",""),risk.get("actual_value",""),risk.get("expected_value",""),crit,risk.get("recommendation","")],1):
                    cell=ws4.cell(row=idx,column=col,value=val); cell.font=NRM; cell.border=BRD; cell.alignment=LFT
                    if fl: cell.fill=fl

    out=io.BytesIO(); wb.save(out); out.seek(0); return out.getvalue()


def render_check(icon, title, detail, css):
    st.markdown(f'<div class="check-item {css}"><div class="check-icon">{icon}</div><div class="check-body"><div class="check-title">{title}</div><div class="check-detail">{detail}</div></div></div>', unsafe_allow_html=True)

def alert(text, kind="info", icon="ℹ️"):
    st.markdown(f'<div class="keden-alert {kind}"><span class="keden-alert-icon">{icon}</span><span>{text}</span></div>', unsafe_allow_html=True)

def section(title, num=""):
    st.markdown(f'<div class="section-header">{num} {title}</div>', unsafe_allow_html=True)

def render_verify(data):
    sanity=data.get("sanity_checks",{}); stats=data.get("statistics",{})
    risks=data.get("risks",[]); cross=data.get("cross_check",{}); permits=data.get("permit_documents",[])

    # Шапка с данными инвойса
    st.markdown(f"""
    <div class="keden-card">
      <div class="keden-card-title"><div class="card-icon">📋</div> Общие сведения по декларации</div>
      <div class="field-grid">
        <div class="field-group"><div class="field-label">Номер инвойса</div><div class="field-value highlight">{stats.get('invoice_number','—')}</div></div>
        <div class="field-group"><div class="field-label">Дата</div><div class="field-value">{stats.get('invoice_date','—')}</div></div>
        <div class="field-group"><div class="field-label">Отправитель</div><div class="field-value">{stats.get('sender','—')}</div></div>
        <div class="field-group"><div class="field-label">Получатель</div><div class="field-value">{stats.get('receiver','—')}</div></div>
        <div class="field-group"><div class="field-label">Страна отправления</div><div class="field-value">{stats.get('country_from','—')}</div></div>
        <div class="field-group"><div class="field-label">Страна назначения</div><div class="field-value">{stats.get('country_to','—')}</div></div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Итоги
    st.markdown(f"""
    <div class="keden-metrics">
      <div class="keden-metric"><div class="keden-metric-value">{stats.get('transit_scenario','—')}</div><div class="keden-metric-label">Сценарий транзита</div></div>
      <div class="keden-metric"><div class="keden-metric-value">{stats.get('total_qty','—')}</div><div class="keden-metric-label">Кол-во (шт)</div></div>
      <div class="keden-metric"><div class="keden-metric-value">{stats.get('total_pkgs','—')}</div><div class="keden-metric-label">Мест (PKGS)</div></div>
      <div class="keden-metric"><div class="keden-metric-value">{stats.get('total_net_weight','—')}</div><div class="keden-metric-label">Нетто (кг)</div></div>
      <div class="keden-metric"><div class="keden-metric-value">{stats.get('total_gross_weight','—')}</div><div class="keden-metric-label">Брутто (кг)</div></div>
      <div class="keden-metric"><div class="keden-metric-value" style="font-size:1.1rem">{stats.get('total_value','—')} {stats.get('currency','')}</div><div class="keden-metric-label">Стоимость</div></div>
      <div class="keden-metric"><div class="keden-metric-value">{stats.get('rows_original','—')}→{stats.get('rows_grouped','—')}</div><div class="keden-metric-label">Строк → HS6 групп</div></div>
    </div>
    """, unsafe_allow_html=True)

    # 1. Проверка данных
    section("Проверка математики и физики данных", "①")
    if sanity.get("passed",True) and not sanity.get("issues"):
        render_check("✅","Все проверки пройдены","Σ Нетто ≤ Σ Брутто · Математика стоимости сходится · Веса и цены положительные","ok")
    else:
        for iss in sanity.get("issues",[]): render_check("⛔","Критическая ошибка",iss,"error")
        alert("Декларацию подавать НЕЛЬЗЯ до исправления данных.","error","⛔")

    # 2. Сверка между документами
    disc=cross.get("discrepancies",[]); matches=cross.get("matches",[]); docs=cross.get("documents_compared",[])
    if docs:
        section("Сверка между документами", "②")
        alert(f"Сравниваются: {' · '.join(docs)}","info","📄")
        if not disc:
            render_check("✅","Расхождений не обнаружено"," · ".join(matches) if matches else "Все ключевые поля совпадают","ok")
        else:
            CM={"КРИТИЧНО":("⛔","error"),"СРЕДНЕ":("⚠️","warning"),"ИНФО":("ℹ️","info")}
            for d in disc:
                ic,cs=CM.get(d.get("criticality","ИНФО"),("ℹ️","info"))
                render_check(ic,f"Расхождение: {d.get('field','—')}",f"Документ 1: {d.get('doc1_value','—')}  ·  Документ 2: {d.get('doc2_value','—')}  ·  {d.get('comment','')}",cs)
            if matches: render_check("✅",f"Совпадают ({len(matches)} полей)"," · ".join(matches),"ok")

    # 3. Замечания
    section("Замечания к качеству данных инвойса", "③")
    if not risks:
        render_check("✅","Замечаний не выявлено","Данные инвойса согласованы","ok")
    else:
        crit_n=sum(1 for r in risks if r.get("criticality")=="КРИТИЧНО")
        med_n=sum(1 for r in risks if r.get("criticality")=="СРЕДНЕ")
        if crit_n: alert(f"Критичных замечаний: {crit_n}  |  Средних: {med_n}","error","⛔")
        else: alert(f"Замечаний: {len(risks)} (критичных нет)","warning","⚠️")
        CM2={"КРИТИЧНО":("⛔","error"),"СРЕДНЕ":("⚠️","warning"),"ИНФО":("ℹ️","info")}
        for r in risks:
            ic,cs=CM2.get(r.get("criticality","ИНФО"),("ℹ️","info"))
            render_check(ic,f"{r.get('problem_type','')} — {r.get('position','')}",f"{r.get('description','')}  ·  Факт: {r.get('actual_value','')}  ·  Ожидалось: {r.get('expected_value','')}  ·  {r.get('recommendation','')}",cs)

    # 4. Разрешительные
    section("Разрешительные документы", "④")
    high=[d for d in permits if d.get("risk_level")=="ВЫСОКИЙ"]
    med=[d for d in permits if d.get("risk_level")=="СРЕДНИЙ"]
    clarif=[d for d in permits if d.get("risk_level")=="ТРЕБУЕТ УТОЧНЕНИЯ"]
    no=[d for d in permits if d.get("risk_level") in ("НЕТ","НИЗКИЙ")]
    if not permits:
        render_check("✅","Ограничений не выявлено","Разрешительные документы не требуются","ok")
    else:
        st.markdown(f"""
        <div class="keden-metrics" style="grid-template-columns:repeat(4,1fr)">
          <div class="keden-metric"><div class="keden-metric-value" style="color:#C0392B">{len(high)}</div><div class="keden-metric-label">🔴 Высокий риск</div></div>
          <div class="keden-metric"><div class="keden-metric-value" style="color:#E67E22">{len(med)}</div><div class="keden-metric-label">🟠 Средний риск</div></div>
          <div class="keden-metric"><div class="keden-metric-value" style="color:#1B5EA8">{len(clarif)}</div><div class="keden-metric-label">🔵 Уточнения</div></div>
          <div class="keden-metric"><div class="keden-metric-value" style="color:#27AE60">{len(no)}</div><div class="keden-metric-label">✅ Без риска</div></div>
        </div>""", unsafe_allow_html=True)
        for d in high: render_check("🔴",f"{d.get('hs6','')} — {d.get('name','')}",f"Акт: {d.get('applicable_act','')}  ·  Документ: {d.get('recommended_doc','')}  ·  Орган: {d.get('issuing_authority','')}","error")
        for d in med: render_check("🟠",f"{d.get('hs6','')} — {d.get('name','')}",f"Акт: {d.get('applicable_act','')}  ·  Документ: {d.get('recommended_doc','')}","warning")
        for d in clarif: render_check("🔵",f"{d.get('hs6','')} — {d.get('name','')}",d.get("comment",""),"info")

    alert("Режим сверки — Excel не создаётся. Для скачивания отчёта выберите другой режим.","info","ℹ️")


# ══════════════════════════════════════════════════════════════════════════════
# ШАПКА
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="keden-header">
  <div class="keden-logo">
    <div class="keden-logo-icon">🛃</div>
    <span>Transit Assist by ИС «KEDEN»</span>
  </div>
  <div class="keden-nav">
    <a href="#">Транзитные декларации</a>
    <a href="#">Помощь</a>
  </div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HERO
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="keden-hero">
  <h1>TRANSIT ASSIST by ИС «KEDEN»</h1>
  <p>Автоматическая обработка инвойсов для транзитных деклараций</p>
  <div class="hero-features">
    <div class="hero-feature">
      <div class="hero-feature-title">Группировка по ТН ВЭД</div>
      <div class="hero-feature-desc">Автоматическое объединение позиций по первым 6 знакам кода</div>
    </div>
    <div class="hero-feature">
      <div class="hero-feature-title">Проверка документов</div>
      <div class="hero-feature-desc">Анализ разрешительных документов для сценариев А и Б</div>
    </div>
    <div class="hero-feature">
      <div class="hero-feature-title">Контроль данных</div>
      <div class="hero-feature-desc">Санити-чеки, сверка весов, математика стоимости</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# ОСНОВНОЙ КОНТЕНТ
# ══════════════════════════════════════════════════════════════════════════════
api_ok = os.getenv("ANTHROPIC_API_KEY","").startswith("sk-ant")

left, right = st.columns([1, 3], gap="small")

with left:
    st.markdown("""
    <div class="keden-sidebar">
      <div style="padding:1rem 1.2rem 0.5rem;border-bottom:1px solid #E0E4EA;margin-bottom:0.5rem">
        <div style="font-size:0.9rem;font-weight:700;color:#1B4F8A">Transit Assist by ИС «KEDEN»</div>
      </div>
      <div class="sidebar-section-title">Обработка инвойса</div>
      <div class="sidebar-item active"><span class="item-icon">📤</span> Загрузка документов</div>
      <div class="sidebar-item"><span class="item-icon">⚙️</span> Выбор режима</div>
      <div class="sidebar-item"><span class="item-icon">📊</span> Результат обработки</div>
      <div class="sidebar-section-title" style="margin-top:1rem">Статус системы</div>
    </div>
    """, unsafe_allow_html=True)
    if api_ok:
        st.markdown('<div style="padding:0 1rem"><div class="keden-alert success"><span>✅</span><span>API ключ подключён</span></div></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div style="padding:0 1rem"><div class="keden-alert error"><span>❌</span><span>API ключ не найден.<br>Проверьте файл .env</span></div></div>', unsafe_allow_html=True)
    st.markdown('<div style="padding:0 1rem;margin-top:0.5rem"><div class="keden-alert info" style="font-size:0.75rem"><span>💰</span><span>~$0.05–0.15 за инвойс</span></div></div>', unsafe_allow_html=True)

with right:
    st.markdown('<div style="padding:1.5rem 0 0 1rem">', unsafe_allow_html=True)

    # Блок загрузки
    st.markdown("""
    <div class="keden-card">
      <div class="keden-card-title"><div class="card-icon">📤</div> Загрузка документов</div>
    </div>
    """, unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Выберите файлы инвойса (Excel или PDF)",
        type=["xlsx","xls","pdf"],
        accept_multiple_files=True,
        help="Можно загрузить несколько файлов: инвойс + упаковочный лист + CMR",
        label_visibility="collapsed"
    )

    if uploaded_files:
        for f in uploaded_files:
            st.markdown(f'<div class="keden-alert success"><span>📄</span><span><b>{f.name}</b> — {round(f.size/1024,1)} КБ</span></div>', unsafe_allow_html=True)

        # Режим
        st.markdown("""
        <div class="keden-card" style="margin-top:1rem">
          <div class="keden-card-title"><div class="card-icon">⚙️</div> Режим обработки</div>
        </div>
        """, unsafe_allow_html=True)

        mode = st.radio("Режим", ["verify","quick","full"],
            format_func=lambda x:{
                "verify":"🔍 Сверка на экране — проверка данных без создания Excel",
                "quick": "⚡ Быстрый режим — только HS6 для KEDEN + Риски (если есть)",
                "full":  "📋 Полный отчёт — 4 листа Excel (Статистика + HS6 + Разрешительные + Риски)",
            }[x], label_visibility="collapsed")

        desc_map = {
            "verify": ('<div class="keden-alert info"><span>🔍</span><span><b>Сверка на экране:</b> Проверка математики · Сверка между документами · Замечания к данным · Разрешительные документы. Excel не создаётся.</span></div>',"info"),
            "quick":  ('<div class="keden-alert warning"><span>⚡</span><span><b>Быстрый режим:</b> Лист HS6 (KEDEN) + Риски при наличии расхождений.</span></div>',"warning"),
            "full":   ('<div class="keden-alert success"><span>📋</span><span><b>Полный отчёт:</b> 4 листа с полным анализом груза.</span></div>',"success"),
        }
        st.markdown(desc_map[mode][0], unsafe_allow_html=True)

        st.markdown("<div style='margin-top:1rem'>", unsafe_allow_html=True)
        run = st.button("🚀 Обработать инвойс", type="primary", use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

        if run:
            prog = st.empty()
            try:
                prog.markdown('<div class="keden-alert info"><span>📂</span><span>Читаем файлы...</span></div>', unsafe_allow_html=True)
                parts=[]
                for uf in uploaded_files:
                    fb=uf.read(); nm=uf.name.lower()
                    if nm.endswith(".pdf"): parts.append(f"=== Файл: {uf.name} (PDF) ===\n{extract_text_from_pdf(fb)}")
                    elif nm.endswith((".xlsx",".xls")): parts.append(f"=== Файл: {uf.name} (Excel) ===\n{extract_text_from_excel(fb)}")
                txt="\n\n".join(parts)
                if len(txt)<50: st.error("Не удалось извлечь текст."); st.stop()

                prog.markdown('<div class="keden-alert info"><span>🤖</span><span>Анализируем в Claude API (30–60 секунд)...</span></div>', unsafe_allow_html=True)
                result=call_claude_api(txt)
                prog.empty()

                sanity=result.get("sanity_checks",{}); stats=result.get("statistics",{})
                if not sanity.get("passed",True):
                    issues="<br>".join(sanity.get("issues",[]))
                    st.markdown(f'<div class="keden-alert error"><span>⛔</span><span><b>СТОП. Критические ошибки:</b><br>{issues}<br>Декларацию подавать НЕЛЬЗЯ.</span></div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div class="keden-alert success"><span>✅</span><span>Sanity checks пройдены: математика сходится, веса корректны.</span></div>', unsafe_allow_html=True)

                st.markdown("""
                <div class="keden-card" style="margin-top:1rem">
                  <div class="keden-card-title"><div class="card-icon">📊</div> Результат обработки</div>
                </div>
                """, unsafe_allow_html=True)

                if mode=="verify":
                    render_verify(result)
                else:
                    prog2=st.empty()
                    prog2.markdown('<div class="keden-alert info"><span>📊</span><span>Формируем Excel-отчёт...</span></div>', unsafe_allow_html=True)
                    excel=create_excel_report(result,mode=mode)
                    prog2.empty()

                    risks=result.get("risks",[]); permits=result.get("permit_documents",[])
                    st.markdown(f"""
                    <div class="keden-metrics">
                      <div class="keden-metric"><div class="keden-metric-value">{stats.get('transit_scenario','—')}</div><div class="keden-metric-label">Сценарий</div></div>
                      <div class="keden-metric"><div class="keden-metric-value" style="font-size:1rem">{stats.get('country_to','—')}</div><div class="keden-metric-label">Страна назначения</div></div>
                      <div class="keden-metric"><div class="keden-metric-value">{stats.get('rows_original','—')}→{stats.get('rows_grouped','—')}</div><div class="keden-metric-label">Строк → HS6</div></div>
                      <div class="keden-metric"><div class="keden-metric-value">{stats.get('total_qty','—')}</div><div class="keden-metric-label">Кол-во (шт)</div></div>
                      <div class="keden-metric"><div class="keden-metric-value">{stats.get('total_net_weight','—')}</div><div class="keden-metric-label">Нетто (кг)</div></div>
                      <div class="keden-metric"><div class="keden-metric-value">{stats.get('total_gross_weight','—')}</div><div class="keden-metric-label">Брутто (кг)</div></div>
                      <div class="keden-metric"><div class="keden-metric-value" style="font-size:1rem">{stats.get('total_value','—')} {stats.get('currency','')}</div><div class="keden-metric-label">Стоимость</div></div>
                    </div>""", unsafe_allow_html=True)

                    if mode=="full":
                        high=sum(1 for d in permits if d.get("risk_level")=="ВЫСОКИЙ")
                        med=sum(1 for d in permits if d.get("risk_level")=="СРЕДНИЙ")
                        cla=sum(1 for d in permits if d.get("risk_level")=="ТРЕБУЕТ УТОЧНЕНИЯ")
                        st.markdown(f"""
                        <div class="keden-metrics" style="grid-template-columns:repeat(4,1fr);margin-top:0.5rem">
                          <div class="keden-metric"><div class="keden-metric-value" style="color:#C0392B">{high}</div><div class="keden-metric-label">🔴 Высокий риск</div></div>
                          <div class="keden-metric"><div class="keden-metric-value" style="color:#E67E22">{med}</div><div class="keden-metric-label">🟠 Средний риск</div></div>
                          <div class="keden-metric"><div class="keden-metric-value" style="color:#1B5EA8">{cla}</div><div class="keden-metric-label">🔵 Уточнения</div></div>
                          <div class="keden-metric"><div class="keden-metric-value" style="color:#566573">{len(risks)}</div><div class="keden-metric-label">⚠️ Замечаний</div></div>
                        </div>""", unsafe_allow_html=True)
                    else:
                        if not risks: alert("Расхождений не обнаружено — лист «Риски» не создавался.","success","✅")
                        else: alert(f"Обнаружено замечаний: {len(risks)} — добавлен лист «Риски и замечания».","warning","⚠️")

                    inv=stats.get("invoice_number","invoice").replace("/","-").replace(" ","_")
                    sfx="ПОЛНЫЙ" if mode=="full" else "HS6"
                    fname=f"KEDEN_{inv}_{sfx}.xlsx"
                    st.markdown('<div style="margin-top:1rem">', unsafe_allow_html=True)
                    st.download_button(
                        f"📥 Скачать Excel-отчёт {'(4 листа)' if mode=='full' else '(1-2 листа)'}",
                        data=excel, file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.markdown("</div>", unsafe_allow_html=True)
                    alert(f"Файл <b>{fname}</b> готов к скачиванию.","success","✅")

            except ValueError as e:
                prog.empty(); st.markdown(f'<div class="keden-alert error"><span>❌</span><span>Ошибка: {e}</span></div>', unsafe_allow_html=True)
            except Exception as e:
                prog.empty(); st.markdown(f'<div class="keden-alert error"><span>❌</span><span>Неожиданная ошибка: {e}</span></div>', unsafe_allow_html=True)
                alert("Проверьте API ключ в файле .env и попробуйте снова.","warning","💡")
    else:
        st.markdown("""
        <div class="keden-card" style="margin-top:1rem; text-align:center; padding:3rem 1.5rem">
          <div style="font-size:3rem;margin-bottom:1rem">📤</div>
          <div style="font-size:1rem;font-weight:600;color:#1B2D45;margin-bottom:0.5rem">Загрузите файл инвойса</div>
          <div style="font-size:0.875rem;color:#6B7C93;line-height:1.6">
            Поддерживаемые форматы: Excel (.xlsx, .xls) и PDF<br>
            Можно загрузить несколько файлов одновременно<br>
            инвойс + упаковочный лист + CMR
          </div>
          <div style="margin-top:1.5rem;display:flex;justify-content:center;gap:1rem;flex-wrap:wrap">
            <div style="background:#EBF3FF;border-radius:6px;padding:0.5rem 1rem;font-size:0.8rem;color:#1B5EA8">🔍 Сверка на экране</div>
            <div style="background:#EBF3FF;border-radius:6px;padding:0.5rem 1rem;font-size:0.8rem;color:#1B5EA8">⚡ Быстрый режим (HS6)</div>
            <div style="background:#EBF3FF;border-radius:6px;padding:0.5rem 1rem;font-size:0.8rem;color:#1B5EA8">📋 Полный отчёт (4 листа)</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# ФУТЕР
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("<div style='height:2rem'></div>", unsafe_allow_html=True)
